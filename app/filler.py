"""Fill the xlsx template with classified items in column B.

Item type -> formatting:
  - name             -> bold + underline, vertical 'center'
  - stage_direction  -> bold,             horizontal 'right',  vertical 'distributed'
  - dialogue         -> regular,          horizontal 'center', vertical 'distributed'

All cells get explicit RTL reading order so mixed Hebrew/English stays
right-to-left regardless of Unicode bidi context.

Row height is computed from text length (approximate wrap calculation) plus
a 10pt padding so the cell never crops.
"""
from __future__ import annotations

import math
import os
from copy import copy
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.pagebreak import Break

TARGET_COLUMN = 2  # column B
DATA_START_ROW = 5
TEMPLATE_LAST_ROW = 125
# Header is rows 1-4. R3 is bumped from the template's 111pt to 174pt by
# _spread_numbered_list_cell (see there for why). Total header occupies the
# top of page 1 and is subtracted from page 1's data budget in the page-break
# walker; page 2+ get the full page_content_height_pt.
# R1=52.5 + R2=42 + R3=174 + R4=56.25 = 324.75pt
HEADER_ROWS_HEIGHT_PT = 324.75
_NUMBERED_LIST_R3_HEIGHT_PT = 174.0

# Print scale (whole percent). Page setup forces this exact scale (rather
# than fit-to-width auto-scale) so the per-page vertical budget below is
# predictable. Calibrated empirically: at narrow margins (0.25" L/R), the
# template's natural column widths (A-M, no widening, no hiding) overflow
# the A4-landscape page width at scale 57+ on the current template —
# every page beyond that gets a phantom right-overflow page tacked on
# (just the M-column 'מכשולים' header + a grid line). The header page
# in particular needs slightly more margin than data pages because the
# merged D2:E3 numbered-list cell and the E4:M4 label band have wider
# effective content. 55% leaves a 1pt safety margin below the empirical
# 56% threshold. If you change this, also update page_content_height_pt
# in _layout_profile (= 487 / scale).
_PRINT_SCALE_PCT = 55

# Per-page vertical budget for content rows. Theoretical max derivation:
# A4 landscape = 595pt tall; minus 0.75" T+B margins (108pt) = 487pt
# physical usable; at _PRINT_SCALE_PCT=61% that's 487 / 0.61 ≈ 798pt of
# logical row-height room per page. Read from _layout_profile() — Carlito
# uses the full 798pt, Noto uses ~5% less to absorb metric drift that was
# clipping multi-line stage_directions at the bottom of pages.

# Empirical layout numbers tuned to the template (Calibri 16, column B width
# 52.71). Hebrew text at this column width and font size wraps at roughly this
# many characters per visual line.
#
# Two-profile setup, gated on CELL_FONT_OVERRIDE:
#   - OFF (local dev, Carlito): the original tuning that produced the
#     reference output. _LINE_HEIGHT_PT=24, paddings small.
#   - ON  (Lambda, Noto Sans Hebrew): Noto renders noticeably taller AND
#     wider than Carlito at the same point size. The OFF tuning undercounts
#     both visual line height and per-line character capacity, which packed
#     rows tight and let long dialogue cells overflow off the bottom of the
#     page (text cut mid-sentence). The ON tuning bumps both line height
#     and per-tier padding, and lowers _CHARS_PER_LINE so wrap estimates
#     run conservative.
_MIN_ROW_HEIGHT_PT = 24.0
_LONG_LINE_THRESHOLD = 4  # >=4 wrapped lines -> "long" padding tier


def _layout_profile() -> dict:
    """Return font-dependent layout constants. Read at call time so changes
    to CELL_FONT_OVERRIDE between requests are picked up without reload.
    """
    if os.environ.get("CELL_FONT_OVERRIDE"):
        return {
            "chars_per_line": 46,
            "line_height_pt": 27.0,
            "name_pad": 6.0,
            "dialogue_short_pad": 6.0,
            "dialogue_long_pad": 14.0,
            "stage_short_pad": 6.0,
            "stage_long_pad": 10.0,
            # Noto undercounts real render height; multi-line stage_directions
            # at the bottom of a page were getting clipped mid-character.
            # ~5% buffer below the theoretical 886pt budget (at 55% scale)
            # eliminates the overflow without losing meaningful rows.
            "page_content_height_pt": 842.0,
        }
    return {
        "chars_per_line": 50,
        "line_height_pt": 24.0,
        "name_pad": 4.0,
        "dialogue_short_pad": 4.0,
        "dialogue_long_pad": 10.0,
        "stage_short_pad": 4.0,
        "stage_long_pad": 6.0,
        # 487pt physical usable / 0.55 scale ≈ 886pt logical per page.
        "page_content_height_pt": 886.0,
    }


def _estimate_row_height(text: str, kind: str) -> float:
    """Estimate the visual row height in points, given the wrapped text.

    Counts explicit newlines plus an approximate wrap based on character
    count. Profile-aware: padding tiers and line height come from
    _layout_profile() so Lambda (Noto) gets looser values than local
    (Carlito).
    """
    if not text:
        return _MIN_ROW_HEIGHT_PT
    p = _layout_profile()
    paragraphs = text.split("\n")
    total_lines = 0
    for para in paragraphs:
        total_lines += max(1, math.ceil(len(para) / p["chars_per_line"]))
    is_long = total_lines >= _LONG_LINE_THRESHOLD
    if kind == "name":
        padding = p["name_pad"]
    elif kind == "dialogue":
        padding = p["dialogue_long_pad"] if is_long else p["dialogue_short_pad"]
    else:  # stage_direction
        padding = p["stage_long_pad"] if is_long else p["stage_short_pad"]
    height = total_lines * p["line_height_pt"] + padding
    return max(_MIN_ROW_HEIGHT_PT, height)


def _alignment_for(kind: str, base: Alignment) -> Alignment:
    if kind == "name":
        h, v = "center", "top"
    elif kind == "stage_direction":
        h, v = "right", "distributed"
    else:  # dialogue
        h, v = "center", "distributed"
    return Alignment(
        horizontal=h,
        vertical=v,
        wrap_text=True,
        readingOrder=2,  # explicit RTL, ignore Unicode bidi base direction
        indent=base.indent or 0,
    )


def _trim_or_extend_rows(ws, last_item_row: int) -> None:
    """Match the table size to the actual item count.

    Uses the worksheet's *actual* max_row rather than a hardcoded constant —
    Excel/openpyxl can leave styled-but-empty cells well past the visible
    data area, which LibreOffice then renders as ghost rows on a trailing
    blank page if we don't remove them.
    """
    actual_last_row = ws.max_row
    if last_item_row < actual_last_row:
        ws.delete_rows(last_item_row + 1, actual_last_row - last_item_row)
    elif last_item_row > actual_last_row:
        src_row = actual_last_row
        for new_row in range(actual_last_row + 1, last_item_row + 1):
            for col in range(1, 6):  # A..E
                src = ws.cell(row=src_row, column=col)
                dst = ws.cell(row=new_row, column=col)
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)


def _set_page_breaks_keeping_name_with_dialogue(ws, items: list[dict]) -> None:
    """Walk the data rows accumulating heights and emit a manual page break
    before every row that would overflow the (conservative) page budget.

    Special case: a name + its following dialogue are treated as one
    indivisible unit, so we never put a manual break between them. If the unit
    doesn't fit on the current page we break *before* the name.

    Because we add manual breaks at every predicted overflow point and the
    page budget is intentionally smaller than LibreOffice's real budget,
    LibreOffice won't insert any natural page breaks of its own — so the page
    break pattern is fully deterministic from this calculation.
    """
    page_budget = _layout_profile()["page_content_height_pt"]
    breaks: list[int] = []  # row numbers AFTER which to break
    current_y = HEADER_ROWS_HEIGHT_PT  # page 1 starts with the header

    i = 0
    while i < len(items):
        item = items[i]
        row = DATA_START_ROW + i
        h = ws.row_dimensions[row].height or _MIN_ROW_HEIGHT_PT

        # If this row is a name and the next is dialogue, plan them as a unit.
        unit_h = h
        unit_size = 1
        if (item["type"] == "name"
                and i + 1 < len(items)
                and items[i + 1]["type"] == "dialogue"):
            next_row = row + 1
            next_h = ws.row_dimensions[next_row].height or _MIN_ROW_HEIGHT_PT
            unit_h = h + next_h
            unit_size = 2

        if current_y + unit_h > page_budget and current_y > HEADER_ROWS_HEIGHT_PT:
            # Doesn't fit on this page -> break before this row.
            breaks.append(row - 1)
            current_y = unit_h
        else:
            current_y += unit_h

        i += unit_size

    if breaks:
        ws.row_breaks.brk = [Break(id=b, man=True) for b in breaks]


def _spread_numbered_list_cell(ws) -> None:
    """The header cell D2 (merged D2:E3) holds an enumerated list rendered
    with vertical=distributed. Excel honors that and spreads the 5 lines
    evenly across the cell; LibreOffice does not, so the lines cluster
    together in the PDF export.

    Fix: insert a blank line between each numbered item so the spacing is
    encoded in the text itself (renderer-independent), switch to
    vertical=top so neither renderer tries to distribute, and bump R3 tall
    enough to fit the now-9-line content. R3 grows by ~63pt; that comes out
    of page 1's data budget (HEADER_ROWS_HEIGHT_PT subtracted from
    _PAGE_CONTENT_HEIGHT_PT in the page-break walker) — about 3 fewer rows
    on page 1. Subsequent pages are unaffected.
    """
    cell = ws.cell(row=2, column=4)  # D2 (merged across D2:E3)
    raw = cell.value
    if not raw or "\n" not in str(raw):
        return
    lines = [ln for ln in str(raw).split("\n") if ln.strip()]
    cell.value = "\n\n".join(lines)
    cell.alignment = Alignment(
        horizontal=cell.alignment.horizontal or "right",
        vertical="top",
        wrap_text=True,
        readingOrder=cell.alignment.readingOrder or 0,
        indent=cell.alignment.indent or 0,
    )
    ws.row_dimensions[3].height = _NUMBERED_LIST_R3_HEIGHT_PT


def _add_breathing_blanks(items: list[dict]) -> list[dict]:
    """Insert blank rows for visual breathing room:

    - Always: one blank row at the very top of the output, so content
      doesn't butt against the header.
    - When the scene opens on one or more stage_directions (i.e. before
      the first name), also insert one blank between the opening stage
      block and the first name. Without this the opening stage paragraph
      runs straight into the first speaker label.

    Blanks are rendered as empty template-styled cells with minimum row
    height (see the kind=='blank' branch in fill_template).
    """
    if not items:
        return items
    out: list[dict] = [{"type": "blank", "text": ""}]
    if items[0]["type"] == "stage_direction":
        i = 0
        while i < len(items) and items[i]["type"] == "stage_direction":
            out.append(items[i])
            i += 1
        if i < len(items):
            out.append({"type": "blank", "text": ""})
        out.extend(items[i:])
    else:
        out.extend(items)
    return out


def fill_template(items: list[dict], template_path: str | Path,
                  output_path: str | Path) -> None:
    items = _add_breathing_blanks(items)
    wb = load_workbook(template_path)
    ws = wb.active

    # CELL_FONT_OVERRIDE workaround for the LibreOffice-on-Lambda Hebrew
    # rendering bug (see comment near the per-item Font assignment below).
    # Apply globally to every cell that already has a font set, so the
    # template's static labels also get readable Hebrew — not just the
    # data cells we fill in below.
    #
    # Also force readingOrder=2 (RTL) on every template cell. The template's
    # static labels rely on font-based bidi inference; once we swap the font
    # to Noto Sans Hebrew that inference shifts, and bidi-neutral chars in
    # mixed Hebrew/ASCII labels drift to the wrong end:
    #   "נסיבות מקדימות:" -> the trailing colon detaches the leading נ
    #   "טקסט + אובייקטים פנימיים + סימון ביטים" -> "+ +" jumps to the front
    # Pinning readingOrder=2 stabilizes the layout regardless of font choice.
    font_override = os.environ.get("CELL_FONT_OVERRIDE")
    if font_override:
        for row in ws.iter_rows():
            for cell in row:
                f = cell.font
                if f is None:
                    continue
                cell.font = Font(
                    name=font_override,
                    size=f.size,
                    bold=f.bold,
                    italic=f.italic,
                    underline=f.underline,
                    color=f.color,
                    strike=f.strike,
                    vertAlign=f.vertAlign,
                )
                a = cell.alignment
                cell.alignment = Alignment(
                    horizontal=a.horizontal,
                    vertical=a.vertical,
                    wrap_text=a.wrap_text,
                    text_rotation=a.text_rotation,
                    indent=a.indent or 0,
                    shrink_to_fit=a.shrink_to_fit,
                    readingOrder=2,
                )

    # Print settings: landscape A4, narrow margins, explicit scale.
    #
    # Why scale=PRINT_SCALE / fitToPage=False (NOT auto-fit-to-width):
    # With fitToPage=True + fitToHeight=0, LibreOffice picks an auto-scale
    # that's much smaller than expected (empirically ~42% in our test files,
    # not the ~66% the column-width math predicts). The result is pages with
    # huge bottom margins — see commit history / CLAUDE.md note about
    # "shrinking to oblivion". Forcing an explicit scale gives predictable
    # output where _PAGE_CONTENT_HEIGHT_PT actually matches the per-page
    # vertical capacity. The template's natural column widths fit the
    # narrow-margin A4-landscape page width at this scale.
    ws.page_setup.orientation = "landscape"
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.sheet_properties.pageSetUpPr.fitToPage = False
    ws.page_setup.fitToWidth = None
    ws.page_setup.fitToHeight = None
    ws.page_setup.scale = _PRINT_SCALE_PCT

    style_cell = ws.cell(row=DATA_START_ROW, column=TARGET_COLUMN)
    base_font = style_cell.font
    base_align = style_cell.alignment

    _spread_numbered_list_cell(ws)

    last_item_row = DATA_START_ROW + len(items) - 1
    _trim_or_extend_rows(ws, last_item_row)

    # Constrain the printable range to A1:M{last_row}. Without this,
    # LibreOffice was emitting a phantom right-overflow page right after
    # the header page (just the M-column 'מכשולים' header band wrapped to
    # its own page). The print_area pins the right edge at column M and
    # the bottom edge at the last filled row so LO has no excuse to make
    # extra pages.
    ws.print_area = f"A1:M{last_item_row}"

    for i, item in enumerate(items):
        row = DATA_START_ROW + i
        kind = item["type"]
        if kind == "blank":
            # Leave the cell with the template's own formatting; just fix the
            # row height so these reserved rows don't balloon vertically.
            ws.row_dimensions[row].height = _MIN_ROW_HEIGHT_PT
            continue
        cell = ws.cell(row=row, column=TARGET_COLUMN)
        bold = kind in ("name", "stage_direction")
        underline = "single" if kind == "name" else None

        cell.value = item["text"]
        # CELL_FONT_OVERRIDE forces a specific font name regardless of the
        # template's setting. Used by the AWS Lambda deploy to work around
        # a LibreOffice bug: when the cell font is Calibri (and Carlito is
        # the substitute), LO picks NotoSansDevanagari as the Hebrew
        # fallback — which has no Hebrew glyphs, so output renders as
        # tofu boxes. Setting CELL_FONT_OVERRIDE="Noto Sans Hebrew" makes
        # Hebrew chars render correctly. Latin chars in the same cells
        # still fall back to Carlito (Calibri-metric) so the page-break
        # math stays approximately correct.
        cell.font = Font(
            name=os.environ.get("CELL_FONT_OVERRIDE") or base_font.name or "Calibri",
            size=base_font.size or 16,
            bold=bold,
            underline=underline,
            color=base_font.color,
        )
        cell.alignment = _alignment_for(kind, base_align)
        ws.row_dimensions[row].height = _estimate_row_height(item["text"], kind)

    _set_page_breaks_keeping_name_with_dialogue(ws, items)

    # openpyxl can leave row_dimensions entries past the filled range
    # (internal bookkeeping during delete_rows / style copies). LibreOffice
    # renders those as ghost rows at default height -> trailing blank pages.
    for r in list(ws.row_dimensions.keys()):
        if r > last_item_row:
            del ws.row_dimensions[r]

    wb.save(output_path)
